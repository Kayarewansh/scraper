"""Reads and writes .xlsx files: the fixed lead-scraper schema, and a
generic reader/writer for arbitrary uploaded spreadsheets (Browse Excel tab,
and its Streamlit equivalent)."""
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("company", "Company Name"),
    ("first_name", "First Name"),
    ("last_name", "Last Name"),
    ("email", "Email"),
    ("phone", "Contact Number"),
    ("address", "Full Address"),
    ("city", "City"),
    ("country", "Country"),
    ("website", "Website"),
    ("review_count", "Google Reviews"),
    ("size_label", "Size Signal (heuristic, not verified revenue)"),
]


def _build_workbook(rows: list, columns: list, sheet_title: str = "Data") -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    for col_idx, header in enumerate(columns, start=1):
        max_len = max([len(str(header))] + [len(str(r.get(header, ""))) for r in rows]) if rows else len(str(header))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return wb


def _leads_rows_and_headers(rows: list):
    renamed_rows = [{header: r.get(key, "") for key, header in COLUMNS} for r in rows]
    headers = [header for _, header in COLUMNS]
    return renamed_rows, headers


def export_to_excel(rows: list, filepath: str) -> None:
    """Writes lead-scraper results using the fixed schema/column labels above."""
    renamed_rows, headers = _leads_rows_and_headers(rows)
    export_generic(renamed_rows, headers, filepath, sheet_title="Leads")


def leads_workbook_bytes(rows: list) -> bytes:
    """Same fixed leads schema as export_to_excel, but returned as in-memory
    .xlsx bytes (for a browser download button instead of a filesystem save)."""
    renamed_rows, headers = _leads_rows_and_headers(rows)
    return generic_workbook_bytes(renamed_rows, headers, sheet_title="Leads")


def export_generic(rows: list, columns: list, filepath: str, sheet_title: str = "Data") -> None:
    """Writes rows (list of dicts) to .xlsx using `columns` as both the
    header labels and the row dict keys, in order — for arbitrary
    uploaded-and-filtered data where there's no separate key/label mapping."""
    wb = _build_workbook(rows, columns, sheet_title)
    wb.save(filepath)


def generic_workbook_bytes(rows: list, columns: list, sheet_title: str = "Data") -> bytes:
    """Same as export_generic, but returns in-memory .xlsx bytes instead of
    saving to a filesystem path."""
    wb = _build_workbook(rows, columns, sheet_title)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_excel(filepath):
    """Reads the first/active sheet of an .xlsx file. `filepath` may be a
    path string or a file-like object (e.g. a Streamlit UploadedFile).
    Returns (columns, rows) — columns is the header row as strings in
    order, rows is a list of dicts keyed by those headers. Blank rows are
    skipped; blank header cells get a placeholder name so every column
    stays addressable."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []

        columns = []
        for i, cell in enumerate(header_row):
            text = str(cell).strip() if cell is not None else ""
            columns.append(text if text else f"Column {i + 1}")

        rows = []
        for row_values in rows_iter:
            if row_values is None or all(v is None for v in row_values):
                continue
            row = {col: ("" if val is None else val) for col, val in zip(columns, row_values)}
            rows.append(row)
        return columns, rows
    finally:
        wb.close()
